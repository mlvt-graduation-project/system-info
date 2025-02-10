import xlsxwriter
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter, column_index_from_string
from datetime import datetime

def CreateExcelTemplate(wb, title="Report Title", sheetName="Sheet1"):
    # Add a worksheet
    if len(wb.sheetnames) == 1 and wb.active.title == "Sheet":
        ws = wb.active
        ws.title = sheetName
    else:
        # ws = wb.add_worksheet(sheetName)
        ws = wb.create_sheet(sheetName)
    
    # Define a bold border style
    thickBorder = Border(
        left=Side(style='thick'), 
        right=Side(style='thick'), 
        top=Side(style='thick'), 
        bottom=Side(style='thick')
    )
    
    # Set the title
    ws.merge_cells('A1:Q1')
    ws['A1'] = title.upper()
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].font = Font(bold=True, size=20)
    # Apply the border to the merged cells
    for row in ws['A1:Q1']:
        for cell in row:
            cell.border = thickBorder
    ws.row_dimensions[1].height = 30   # Set the row height for the title

    # Set the report date
    ws.merge_cells('A2:Q2')
    reportDate = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    reportTime = "Report Time: " + reportDate
    ws['A2'] = reportTime
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A2'].font = Font(bold=True, size=14)
    # Apply the border to the merged cells
    for row in ws['A2:Q2']:
        for cell in row:
            cell.border = thickBorder
    ws.row_dimensions[2].height = 20  # Set the row height for the report time
    
    # # Optionally set column widths
    # worksheet.set_column('A:A', 20)  
    # worksheet.set_column('B:B', 20)  

    return ws  # Return the worksheet for further use

def ExportExcel(listColumn, listRow, fileName, sheetTitle):
    # Create a workbook
    # workbook = xlsxwriter.Workbook(fileName)
    wb = Workbook()
    
    # Call create_excel_template to initialize the template
    worksheet = CreateExcelTemplate(wb, title=sheetTitle)
    
    # Define the cell border style
    thinBorder = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    # Write the headers from listColumn
    headers = listColumn.split(', ')
    for colNum, header in enumerate(listColumn.split(', ')):
        cell = worksheet.cell(row=4, column=colNum+1, value=header.strip())  # Start writing headers from row 2
        cell.border = thinBorder

    # Write the data from listRow
    columnWidths = [0] * len(headers)
    for rowNum, row in enumerate(listRow, 4):  # Start writing data from row 3
        for colNum, item in enumerate(row):
            cell = worksheet.cell(row=rowNum, column=colNum+1, value=item)
            cell.border = thinBorder
            if len(str(item)) > columnWidths[colNum]:
                columnWidths[colNum] = len(str(item))
                
    # Adjust column widths
    for i, column_width in enumerate(columnWidths, start=1):
        worksheet.column_dimensions[get_column_letter(i)].width = column_width + 2  # Add 2 for padding

    # Close the workbook
    wb.save(fileName)
    wb.close()

def ExportExcelWithChartImage(fileName, sheetName, data, imagePath, title, chartType='pie', chartName='Chart'):
    # Create a DataFrame from the data
    df = pd.DataFrame(data)
    
    # Create a new workbook and select the active worksheet
    wb = Workbook()
    ws = CreateExcelTemplate(wb, title=title, sheetName=sheetName)
    
    # Populate the worksheet with DataFrame's data and add border to each cell
    thinBorder = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    columnWidths = {}
    for r_idx, r in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
        for c_idx, value in enumerate(r, start=1):
            cell = ws.cell(row=r_idx + 3, column=c_idx, value=value)
            cell.border = thinBorder
            if len(str(value)) > columnWidths.get(c_idx, 0):
                columnWidths[c_idx] = len(str(value))

    # Adjust column widths
    for col_idx, width in columnWidths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width + 2
    
    # # Populate the worksheet with DataFrame's data
    # for r in dataframe_to_rows(df, index=False, header=True):
    #     ws.append(r)
    
    # Create chart based on the specified type
    plt.figure(figsize=(8, 8))
    labels = df.iloc[:, 0].tolist()  
    values = df.iloc[:, 1].tolist()  

    if chartType == 'pie':
        def autopct_format(pct):
            total = sum(values)
            val = int(round(pct*total/100.0))
            for i, value in enumerate(values):
                if value == val:
                    return '{v}\n({p:.1f}%)'.format(v=labels[i], p=pct)
            return ''
        
        wedges, texts, autotexts = plt.pie(values, labels=labels, autopct=autopct_format, startangle=140, 
                wedgeprops={'edgecolor': 'white'}, textprops={'color': 'white'})
        
        # Customize and position the legend below the chart
        plt.legend(wedges, labels, loc="upper center", bbox_to_anchor=(0.5, -0.05), 
                   fancybox=True, shadow=True, ncol=len(labels) // 2)
        
        plt.title(chartName)
    elif chartType == 'bar':
        plt.bar(labels, values, color='lightblue')
        plt.title(chartName)
        plt.xticks(rotation=45)
        plt.grid(True, linestyle='-', which='major', color='lightgrey', alpha=0.5)  # Add grid lines
        plt.gca().axes.get_yaxis().set_visible(True)  # Keep the Y-axis visible
        plt.gca().axes.get_xaxis().set_tick_params(labelsize=10)  # Customize the X-axis tick label size
        plt.gca().axes.get_xaxis().set_visible(True)  # Ensure X-axis is visible but without specific title
    plt.axis('equal') if chartType == 'pie' else plt.tight_layout()
    plt.savefig(imagePath)  # Save the chart as an image
    plt.close()

    # Determine the image placement column dynamically
    data_length = len(data)
    base_column = 'E'  # Starting point if the list is shorter or equal to 2
    if data_length > 2:
        column_index = column_index_from_string(base_column) + (data_length - 2)  # Adjust starting from 'E'
        new_column_letter = get_column_letter(column_index)
    else:
        new_column_letter = base_column

    # Load and add the image to the worksheet
    img = Image(imagePath)
    img_cell = f'{new_column_letter}4'  # Assuming you want it at row 4
    ws.add_image(img, img_cell) # Adjust position based on the data rows

    # Save the workbook
    wb.save(fileName)

# Example usage:
listColumn = "Name, Age, City"
listRow = [("Alice", 30, "New York"), ("Bob", 25, "Los Angeles"), ("Charlie", 35, "Chicago")]

# Call the function to export data
ExportExcel(listColumn, listRow, "Report.xlsx", "Monthly Sales Report")

# Data for the charts
data = {
    'Brand': ['Samsung', 'Apple', 'Xiaomi', 'Oppo', 'Vivo', 'Transsion', 'Honor', 'Rest'],
    'Market Share': [21, 17, 14, 9, 9, 10, 6, 8]
}

# Call the function for a pie chart
ExportExcelWithChartImage("Excel_with_Pie_Chart_Image.xlsx", "Data", data, "pie_chart.png", "Report", 'pie', 'Pie Chart')

# Call the function for a bar chart
ExportExcelWithChartImage("Excel_with_Bar_Chart_Image.xlsx", "Data", data, "bar_chart.png", "Report", 'bar', 'Bar Chart')